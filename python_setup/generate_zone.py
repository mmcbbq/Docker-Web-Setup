import openpyxl


def generate_zone(domain, excel_file, network, ip):
    wb_obj_read = openpyxl.load_workbook(excel_file)
    sheet_read = wb_obj_read.active
    dns_zone = open(f'../bind/{domain}.fwd', 'w')
    dns_zone.write(f'''$TTL	86400
@	IN	SOA	ns-1.{domain}. root.{domain}.zz. (
			      1		; Serial
			 604800		; Refresh
			  86400		; Retry
			2419200		; Expire
			  86400 )	; Negative Cache TTL
;



@		IN	NS	ns-1.{domain}.
ns-1		IN	A	{network}.{ip}
www.manuel.{domain}.		IN	A	{network}.{ip}
''')

    max_col = sheet_read.max_column
    max_row = sheet_read.max_row

    for x in range(2, max_row):
        dns_zone.write(
            f'{sheet_read.cell(row=x, column=3).value}. \t IN \t A {sheet_read.cell(row=x, column=4).value} \n')
