import openpyxl


def genExcel(excelfile, domain, ip_start, ip_step, network):
    wb_obj = openpyxl.load_workbook(excelfile)
    sheet = wb_obj.active

    # sheet_obj = wb_obj.active

    max_col = sheet.max_column
    max_row = sheet.max_row

    teil = []

    for j in range(2, max_row + 1):
        sub = sheet.cell(row=j, column=2).value
        c = 1
        sub = sub.replace(' ', '')
        for i,k in {'ö': 'o', 'ä': 'a', 'ü': 'u'}.items():
            # print('check', i, 'in' ,sub)
            if i in sub:
                sub = sub.replace(i, f'{k}e')
        # while sub in teil:
        #     sub = sheet.cell(row=j, column=1).value[0 + 1:2 + 1] + sheet.cell(row=j, column=2).value[0]
        #     c += 1
        # teil.append(sub.lower())
        sheet.cell(row=j, column=3, value=f'www.{sub.lower()}-{domain}')
        sheet.cell(row=j, column=4, value=f'{network}.{ip_start}')
        ip_start += ip_step
    wb_obj.save(excelfile)
